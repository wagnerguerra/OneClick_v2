"""E2E do cli.py: as abas D101/D105 saem no XLSX com cabeçalho e vínculo ao D100."""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

ENGINE_DIR = Path(__file__).resolve().parents[1] / "sped_engine"
FIXTURE = Path(__file__).parent / "fixtures" / "sped_bloco_d.txt"


def _run_cli(out_path: Path, *extra: str):
    proc = subprocess.run(
        [sys.executable, str(ENGINE_DIR / "cli.py"), "--input", str(FIXTURE), "--output", str(out_path), *extra],
        cwd=str(ENGINE_DIR),
        capture_output=True,
        text=True,
    )
    assert proc.returncode == 0, proc.stderr or proc.stdout
    linhas = [json.loads(l) for l in proc.stdout.splitlines() if l.startswith("{")]
    assert any(x.get("kind") == "done" for x in linhas), proc.stdout
    assert not any(x.get("kind") == "error" for x in linhas), proc.stdout
    return out_path


def test_export_seleciona_apenas_d101_d105(tmp_path):
    out = _run_cli(tmp_path / "d101_d105.xlsx", "--sheets", "D101,D105")
    wb = openpyxl.load_workbook(out, read_only=True)
    assert wb.sheetnames == ["D101", "D105"]

    ws = wb["D101"]
    linhas = list(ws.iter_rows(values_only=True))
    cabecalho = [c for c in linhas[0] if c is not None]
    assert "NUM_DOC" in cabecalho and "CHV_CTE" in cabecalho and "VL_PIS" in cabecalho
    idx_num = cabecalho.index("NUM_DOC")
    assert len(linhas) - 1 == 2  # duas linhas de dados
    assert str(linhas[1][idx_num]) == "12345"
    assert str(linhas[2][idx_num]) == "67890"
    wb.close()


def test_export_completo_inclui_d101_d105(tmp_path):
    out = _run_cli(tmp_path / "full.xlsx")
    wb = openpyxl.load_workbook(out, read_only=True)
    nomes = wb.sheetnames
    wb.close()
    assert "D101" in nomes and "D105" in nomes
    assert nomes.index("D100") < nomes.index("D101") < nomes.index("D190")
