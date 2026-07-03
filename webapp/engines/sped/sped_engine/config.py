from openpyxl.styles import PatternFill, Font

# SHEET_ORDER: manter igual a SPED_EXPORT_SHEET_KEYS em webapp-01/packages/contracts/src/index.ts

HEADERS = {
    "0150": ["REG","COD_PART","NOME","COD_PAIS","CNPJ","CPF","IE","COD_MUN","SUFRAMA","END","NUM","COMPL","BAIRRO"],
    "0200": ["REG","COD_ITEM","DESCR_ITEM","COD_BARRA","COD_ANT_ITEM","UNID_INV","TIPO_ITEM","COD_NCM","EX_IPI",
             "COD_GEN","COD_LST","ALIQ_ICMS","CEST"],
    "C100": ["REG","IND_OPER","IND_EMIT","COD_PART","COD_MOD","COD_SIT","SER","NUM_DOC","CHV_NFE","DT_DOC","DT_E_S",
             "VL_DOC","IND_PGTO","VL_DESC","VL_ABAT_NT","VL_MERC","IND_FRT","VL_FRT","VL_SEG","VL_OUT_DA",
             "VL_BC_ICMS","VL_ICMS","VL_BC_ICMS_ST","VL_ICMS_ST","VL_IPI","VL_PIS","VL_COFINS","VL_PIS_ST","VL_COFINS_ST"],
    "C170": ["REG","NUM_DOC","CHV_NFE","NUM_ITEM","COD_ITEM","DESCR_COMPL","QTD","UNID","VL_ITEM","VL_DESC","IND_MOV","CST_ICMS","CFOP","COD_NAT",
             "VL_BC_ICMS","ALIQ_ICMS","VL_ICMS","VL_BC_ICMS_ST","ALIQ_ST","VL_ICMS_ST","IND_APUR","CST_IPI","COD_ENQ",
             "VL_BC_IPI","ALIQ_IPI","VL_IPI","CST_PIS","VL_BC_PIS","ALIQ_PIS","QUANT_BC_PIS","ALIQ_PIS_R","VL_PIS",
             "CST_COFINS","VL_BC_COFINS","ALIQ_COFINS","QUANT_BC_COFINS","ALIQ_COFINS_R","VL_COFINS","COD_CTA","VL_ABAT_NT"],
    "C190": ["REG","NUM_DOC","CHV_NFE","CST_ICMS","CFOP","ALIQ_ICMS","VL_OPR","VL_BC_ICMS","VL_ICMS","VL_BC_ICMS_ST","VL_ICMS_ST","VL_RED_BC","VL_IPI","COD_OBS"],
    "C500": ["REG","IND_OPER","IND_EMIT","COD_PART","COD_MOD","COD_SIT","SER","SUB","COD_CONS","NUM_DOC","DT_DOC","DT_E_S",
             "VL_DOC","VL_DESC","VL_FORN","VL_SERV_NT","VL_TERC","VL_DA","VL_BC_ICMS","VL_ICMS","VL_BC_ICMS_ST","VL_ICMS_ST",
             "COD_INF","VL_PIS","VL_COFINS","TP_LIGACAO","COD_GRUPO_TENSAO"],
    "C590": ["REG","NUM_DOC","CST_ICMS","CFOP","ALIQ_ICMS","VL_OPR","VL_BC_ICMS","VL_ICMS","VL_BC_ICMS_ST","VL_ICMS_ST","VL_RED_BC","COD_OBS"],
    "D100": ["REG","IND_OPER","IND_EMIT","COD_PART","COD_MOD","COD_SIT","SER","SUB","NUM_DOC","CHV_CTE","DT_DOC","DT_A_P",
             "TP_CTE","CHV_CTE_REF","VL_DOC","VL_DESC","IND_FRT","VL_SERV","VL_BC_ICMS","VL_ICMS","VL_NT","COD_INF",
             "COD_CTA","COD_MUN_ORIG","COD_MUN_DEST"],
    "D190": ["REG","NUM_DOC","CHV_CTE","CST_ICMS","CFOP","ALIQ_ICMS","VL_OPR","VL_BC_ICMS","VL_ICMS","VL_RED_BC","COD_OBS"],
    "D500": ["REG","IND_OPER","IND_EMIT","COD_PART","COD_MOD","COD_SIT","SER","SUB","NUM_DOC","DT_DOC","DT_A_P","VL_DOC",
             "VL_DESC","VL_SERV","VL_SERV_NT","VL_TERC","VL_DA","VL_BC_ICMS","VL_ICMS","COD_INF","VL_PIS","VL_COFINS","TP_ASSINANTE"],
    "D590": ["REG","NUM_DOC","CST_ICMS","CFOP","ALIQ_ICMS","VL_OPR","VL_BC_ICMS","VL_ICMS","VL_BC_ICMS_UF","VL_ICMS_UF","VL_RED_BC","COD_OBS"],
    # REG 0450 — informação complementar do documento fiscal (guia cabecalhos_sped.txt)
    "0450": ["REG", "COD_INF", "TXT"],
    # REG K200 — Bloco K, estoque escriturado (Guia EFD ICMS/IPI)
    "K200": ["REG", "DT_EST", "COD_ITEM", "QTD", "IND_EST", "COD_PART"],
}
# Ordem canónica dos 11 blocos principais — manter igual a SPED_EXPORT_SHEET_KEYS nos contracts
SHEET_ORDER = [
    "0150",
    "0200",
    "C100",
    "C170",
    "C190",
    "C500",
    "C590",
    "D100",
    "D190",
    "D500",
    "D590",
]

HEADER_FILL  = PatternFill(start_color="FF4169E1", end_color="FF4169E1", fill_type="solid")
ALT_FILL     = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFFFF", bold=True)