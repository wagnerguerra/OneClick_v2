from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font
from datetime import datetime

# Paleta de cores
HEADER_FILL = PatternFill(start_color="FF4169E1", end_color="FF4169E1", fill_type="solid")  # azul cabeçalho (padrão 4169E1)
ALT_FILL    = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")  # zebra
SECTION_FILL= PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")  # cinza claro

HEADER_FONT = Font(color="FFFFFFFF", bold=True)
SECTION_FONT= Font(color="000000", bold=True, size=12)

HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN   = Alignment(horizontal="center", vertical="center")

class DefaultReportBuilder:

    def _format_table(self, ws, start_row, end_row):
        """Aplica formatação para uma tabela (cabeçalho + zebra + alinhamento)."""
        if start_row > end_row:
            return

        # Cabeçalho da tabela (linhas 6, 21, etc.)
        for c in ws[start_row]:
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # Linhas de dados
        for r in range(start_row + 1, end_row + 1):
            if (r - start_row) % 2 == 1:
                for c in ws[r]:
                    c.fill = ALT_FILL
            for c in ws[r]:
                if isinstance(c.value, (int, float)):
                    c.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

        # Filtro automático
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(ws.max_column)}{end_row}"


    def _format_section_title(self, ws, row, text):
        """Insere título de seção formatado."""
        ws.append([text])
        cell = ws[f"A{row}"]
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    def write_report(
        self,
        path,
        summary,
        mismatches,
        link_checks=None,
        razao=None,
        cnpj=None,
        generic_export_regs=None,
    ):
        wb = load_workbook(path)
        if "RELATORIO" in wb.sheetnames:
            del wb["RELATORIO"]
        ws = wb.create_sheet("RELATORIO")

        # === Cabeçalho do relatório ===
        ws.append([f"Empresa: {razao or 'N/D'}"])
        ws.append([f"CNPJ: {cnpj or 'N/D'}"])
        ws.append([f"Data de geração: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"])
        ws.append([])

        for row in range(1, 4):
            ws[f"A{row}"].font = Font(bold=True, size=12)
            ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            ws.row_dimensions[row].height = 20

        # === Resumo dos registros ===
        self._format_section_title(ws, ws.max_row + 1, "📊 Resumo dos Registros")
        ws.append(["REGISTRO", "QUANTIDADE", "COLUNAS EXTRAS", "REGISTROS FORA DO PADRÃO"])
        start_row = ws.max_row
        for r in summary:
            ws.append([r["REGISTRO"], r["LINHAS"], r["MAX_EXTRAS"], r["MISMATCH_REG"]])
        self._format_table(ws, start_row, ws.max_row)

        # Dentro de DefaultReportBuilder.write_report

        # === Validação de vínculos ===
        ws.append([])
        self._format_section_title(ws, ws.max_row + 1, "🔗 Validação de Vínculos")
        ws.append(["VÍNCULO", "LINHAS NO EXCEL", "LINHAS NO TXT", "POSSUI NUM_DOC", "POSSUI CHAVE", "DIVERGÊNCIAS"])
        start_row = ws.max_row
        if link_checks:
            for k in ("C170", "C190", "C590", "D190", "D590"):
                info = link_checks.get(k, {})
                if not info:
                    continue
                if k in ("C170", "C190"):
                    vinculo = "C100 → " + k
                elif k == "C590":
                    vinculo = "C500 → C590"
                elif k == "D190":
                    vinculo = "D100 → D190"
                else:  # D590
                    vinculo = "D500 → D590"
                ws.append([
                    vinculo,
                    info.get("rows_excel"),
                    info.get("rows_expected"),
                    info.get("has_NUM_DOC"),
                    info.get("has_CHV_NFE") or info.get("has_CHV_CTE") or info.get("has_CHV"),
                    info.get("mismatch_count"),
                ])
        self._format_table(ws, start_row, ws.max_row)

        if generic_export_regs:
            ws.append([])
            self._format_section_title(ws, ws.max_row + 1, "Registros exportados com layout genérico (COL_xx)")
            start_row = ws.max_row + 1
            ws.append(["REGISTRO", "NOTA"])
            for reg in generic_export_regs:
                ws.append([
                    reg,
                    "Sem linha em cabecalhos_sped.txt — colunas COL_01, COL_02, … na ordem do arquivo",
                ])
            self._format_table(ws, start_row, ws.max_row)

        # === Divergências detalhadas ===
        ws.append([])
        self._format_section_title(ws, ws.max_row + 1, "⚠️ Divergências Detalhadas")
        if link_checks:
            for k in ("C170", "C190", "C590", "D190", "D590"):
                info = link_checks.get(k, {})
                if not info or not info.get("mismatches"):
                    continue
                ws.append([])
                ws.append([f"Registro {k} - divergências (amostra limitada a 200)"])
                start_row = ws.max_row + 1
                ws.append(["ÍNDICE", "NUM_DOC (esperado)", "NUM_DOC (obtido)",
                           "CHAVE (esperada)", "CHAVE (obtida)"])
                for mm in info["mismatches"][:200]:
                    ws.append([
                        mm["index"],
                        mm.get("expected_NUM_DOC"),
                        mm.get("got_NUM_DOC"),
                        mm.get("expected_CHV_NFE") or mm.get("expected_CHV_CTE"),
                        mm.get("got_CHV_NFE") or mm.get("got_CHV_CTE"),
                    ])
                self._format_table(ws, start_row, ws.max_row)


        # === Registros fora do padrão ===
        ws.append([])
        self._format_section_title(ws, ws.max_row + 1, "🚨 Registros fora do padrão estrutural")
        start_row = ws.max_row + 1
        ws.append(["REGISTRO", "LINHA", "VALOR ENCONTRADO NO CAMPO REG"])
        for r in mismatches:
            ws.append([r["REGISTRO"], r["LINHA_IDX"], r["VALOR_REG_ENCONTRADO"]])
        self._format_table(ws, start_row, ws.max_row)

        # === Ajuste global da aba ===
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # Altura uniforme
        for r in range(1, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        # Ajuste automático da largura das colunas
        for col in ws.columns:
            maxlen = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(120, max(10, maxlen + 2))

        # Congelar cabeçalho até as seções
        ws.freeze_panes = "A6"

        # Mover aba RELATORIO para primeira posição
        ws_rel = wb["RELATORIO"]
        wb._sheets.remove(ws_rel)
        wb._sheets.insert(0, ws_rel)

        wb.save(path)
