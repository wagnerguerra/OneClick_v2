from __future__ import annotations
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import unicodedata
import re

# ==============================
# Configurações visuais gerais
# ==============================
HEADER_FONT   = Font(bold=True)
HEADER_HEIGHT = 28
ROW_HEIGHT    = 22

# Cores (ARGB)
HEADER_FILL = PatternFill(start_color="FF4169E1", end_color="FF4169E1", fill_type="solid")  # 4169E1 azul padrão
ROW_FILL    = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")  # DCE6F1 azul claro

# ==============================
# Funções auxiliares
# ==============================
def _auto_width(ws: Worksheet) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[letter]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[letter].width = min(max(8, max_len + 2), 60)

def _center_all(ws: Worksheet) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str):
                cell.value = cell.value.replace("\\n", " ").replace("\\r", " ")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def _header_and_row_heights(ws: Worksheet) -> None:
    if ws.max_row >= 1:
        ws.row_dimensions[1].height = HEADER_HEIGHT
        for c in ws[1]:
            if isinstance(c, MergedCell):
                continue
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
    for r in range(2, (ws.max_row or 1) + 1):
        ws.row_dimensions[r].height = ROW_HEIGHT

def _highlight_delta_headers(ws: Worksheet) -> None:
    """Destaca cabeçalhos que começam com 'Delta' (case-sensitive)."""
    fill = PatternFill(start_color="FFFDE9D9", end_color="FFFDE9D9", fill_type="solid")
    if ws.max_row >= 1:
        for cell in ws[1]:
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.strip().startswith("Delta"):
                cell.fill = fill

def _borders_and_rows(ws: Worksheet) -> None:
    """Borda fina + azul claro uniforme em todas as linhas de dados."""
    thin   = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    max_r, max_c = ws.max_row, ws.max_column
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c), start=1):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.border = border
            if r_idx > 1:
                cell.fill = ROW_FILL


def _space_items_xml(ws: Worksheet) -> None:
    """Insere 2 linhas em branco quando a CHAVE mudar (para 'Itens (XML)')."""
    if ws.max_row < 3:
        return
    header = [c.value for c in ws[1]]
    try:
        chave_col = header.index("CHAVE") + 1  # 1-based
    except ValueError:
        return
    r = ws.max_row
    while r > 2:
        atual = ws.cell(row=r,   column=chave_col).value
        acima = ws.cell(row=r-1, column=chave_col).value
        if atual and acima and atual != acima:
            ws.insert_rows(r, amount=2)
        r -= 1





# -----------------------------
# Coerção numérica por cabeçalho
# -----------------------------
def _norm_header(txt: str) -> str:
    """Remove acentos, deixa UPPER, tira sufixos ' (SCI)'/' (Cliente)' para comparar."""
    if txt is None:
        return ""
    s = str(txt)
    # remove sufixo entre parênteses: "Vlr contábil (SCI)" -> "Vlr contábil"
    if " (" in s:
        s = s.split(" (", 1)[0]
    # normaliza acentos e caixa
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s.upper().strip()

_NUMERIC_FLOAT_HEADERS = {
    "VLR CONTABIL", "VALOR CONTABIL", "VLR CONTÁBIL", "VALOR CONTÁBIL",
    "BASE DE ICMS", "BASE ICMS",
    "VALOR DO ICMS", "VALOR ICMS",
    "BASE DE IPI", "BASE IPI",
    "VALOR DO IPI", "VALOR IPI",
    # variações com "SOMA DE ..."
    "SOMA DE VLR. CONTABIL", "SOMA DE VLR. CONTÁBIL",
    "SOMA DE BASE ICMS",
    "SOMA DE VLR. ICMS",
    "SOMA DE BASE IPI",
    "SOMA DE VLR. IPI",
}

_NUMERIC_INT_HEADERS = {"NOTA", "Nº NF.", "Nº NF", "NO NF", "NUM NF", "NUMERO NF", "NF"}

def _parse_float_like(v):
    """Converte strings com R$, ponto de milhar e vírgula decimal em float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # remove moeda e espaços
    s = s.replace("R$", "").replace("\u00A0", " ").replace(" ", "")
    # troca separadores: 1.234.567,89 -> 1234567.89
    s = s.replace(".", "").replace(",", ".")
    # mantém apenas dígitos, ponto e sinal
    s = re.sub(r"[^0-9\.\-]", "", s)
    try:
        return float(s) if s not in ("", "-", ".", "-.") else None
    except Exception:
        return None

def _parse_int_like(v):
    """Extrai dígitos de algo tipo '001234' e devolve int (ou None se não houver)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return int(round(float(v)))
        except Exception:
            return None
    m = re.search(r"(\d+)", str(v))
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None

def _coerce_numeric_by_header(ws: Worksheet) -> None:
    """Converte texto->número nas colunas reconhecidas e aplica number_format."""
    if ws.max_row < 2 or ws.max_column < 1:
        return

    # mapeia colunas-alvo pelo cabeçalho
    header_vals = [c.value for c in ws[1]]
    targets = {}  # col_idx -> ("int"|"float")
    for idx, h in enumerate(header_vals, start=1):
        nh = _norm_header(h)
        if nh in _NUMERIC_FLOAT_HEADERS:
            targets[idx] = "float"
        elif nh in _NUMERIC_INT_HEADERS:
            targets[idx] = "int"

    if not targets:
        return

    # percorre linhas e converte valor + aplica formato
    for r in range(2, ws.max_row + 1):
        for c_idx, kind in targets.items():
            cell = ws.cell(row=r, column=c_idx)
            # ignora se já for número
            if isinstance(cell.value, (int, float)):
                # só garante o formato
                if kind == "float":
                    cell.number_format = "#,##0.00"
                else:
                    cell.number_format = "0"
                continue

            if kind == "float":
                val = _parse_float_like(cell.value)
                if val is not None:
                    cell.value = float(val)
                    cell.number_format = "#,##0.00"
            else:  # int
                val = _parse_int_like(cell.value)
                if val is not None:
                    cell.value = int(val)
                    cell.number_format = "0"

# ==============================
# Orquestração principal
# ==============================
def format_workbook(wb: Workbook) -> None:
    """Aplica formatação padrão em todas as abas + regras específicas."""
    for ws in wb.worksheets:
        _header_and_row_heights(ws)
        _coerce_numeric_by_header(ws)
        _center_all(ws)
        _highlight_delta_headers(ws)
        _auto_width(ws)
        _borders_and_rows(ws)

    # Espaçamento específico para a aba de itens XML (se existir)
    if "Itens (XML)" in wb.sheetnames:
        _space_items_xml(wb["Itens (XML)"])
