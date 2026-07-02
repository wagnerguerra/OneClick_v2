#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parent
_SPED_ENGINE = _ROOT.parent / "sped" / "sped_engine"
if _SPED_ENGINE.is_dir():
    sys.path.insert(0, str(_SPED_ENGINE))

from cabecalhos_sped import merge_headers  # noqa: E402
from config import HEADERS  # noqa: E402

CORE_SHEETS = ("0150", "0200", "C100", "C170", "C190", "C500", "C590", "D100", "D190", "D500", "D590")
REG_RE = re.compile(r"^[0-9A-Z]{4}$")
MERGE_HEADERS = merge_headers(HEADERS)


def inspect_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    reasons: list[str] = []
    reg_sheets: list[str] = []
    line_numbers: set[int] = set()

    for name in wb.sheetnames:
        n = str(name).strip().upper()
        if not REG_RE.fullmatch(n):
            continue
        reg_sheets.append(n)
        ws = wb[name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(x or "").strip() for x in (header_row or [])]
        if "_LINHA" not in headers:
            reasons.append(f"Aba {n}: coluna _LINHA ausente")
            continue
        if n in MERGE_HEADERS:
            expected_cols = MERGE_HEADERS[n]
            missing_cols = [c for c in expected_cols if c not in headers]
            if missing_cols:
                reasons.append(f"Aba {n}: colunas obrigatórias ausentes ({','.join(missing_cols[:20])})")
                continue
        line_col = headers.index("_LINHA")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if line_col >= len(row):
                continue
            raw = row[line_col]
            if raw in (None, ""):
                continue
            try:
                ln = int(float(raw))
            except (TypeError, ValueError):
                reasons.append(f"Aba {n}: _LINHA inválida ({raw})")
                continue
            if ln < 1:
                reasons.append(f"Aba {n}: _LINHA inválida ({ln})")
                continue
            line_numbers.add(ln)

    missing_core = [s for s in CORE_SHEETS if s not in reg_sheets]
    if missing_core:
        reasons.append("Abas core ausentes: " + ",".join(missing_core))

    if line_numbers:
        mx = max(line_numbers)
        expected = set(range(1, mx + 1))
        gaps = sorted(expected - line_numbers)
        if gaps:
            preview = ",".join(str(x) for x in gaps[:20])
            reasons.append(f"Linhas _LINHA faltando na sequência 1..{mx} (ex.: {preview})")
    else:
        reasons.append("Nenhum _LINHA válido encontrado")

    complete = len(reasons) == 0
    return {
        "complete": complete,
        "requiresOriginal": not complete,
        "reasons": reasons,
        "regSheets": sorted(set(reg_sheets)),
    }


def main() -> int:
    p = argparse.ArgumentParser(description="Inspeciona completude do XLSX para modo sem SPED original")
    p.add_argument("--xlsx", required=True, help="Planilha de entrada")
    args = p.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.is_file():
        print(json.dumps({"kind": "error", "message": f"XLSX não encontrado: {xlsx}"}), flush=True)
        return 1
    try:
        out = inspect_xlsx(xlsx)
        print(json.dumps({"kind": "ok", **out}), flush=True)
        return 0
    except Exception as e:
        print(json.dumps({"kind": "error", "message": str(e)}), flush=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())

