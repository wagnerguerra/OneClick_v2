"""Gera XLSX formatado com 2 abas: Lançamentos (sucessos) e Falhas (erros)."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11, color="1A1A1F")

_thin = Side(border_style="thin", color="CECECE")
CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

CURRENCY_FMT = 'R$ #,##0.00;[Red]-R$ #,##0.00'

LANC_COLUMNS = [
    ("Arquivo", "arquivo", "text"),
    ("CNPJ/CPF Destinatário", "cnpj_destinatario", "cpf_cnpj"),
    ("Valor Principal", "valor_principal", "currency"),
    ("UF", "uf_favorecida", "text"),
    ("Vencimento", "data_vencimento", "date_br"),
    ("Período", "periodo_referencia", "text"),
    ("Nº Controle", "no_controle", "text"),
    ("Registrado em", "criado_em", "text"),
]

FALHAS_COLUMNS = [
    ("Arquivo", "arquivo", "text"),
    ("Motivo", "motivo", "text_wrap"),
    ("Registrado em", "criado_em", "text"),
]


def _fmt_cpf_cnpj(d: str | None) -> str:
    if not d:
        return ""
    s = "".join(c for c in d if c.isdigit())
    if len(s) == 11:
        return f"{s[:3]}.{s[3:6]}.{s[6:9]}-{s[9:11]}"
    if len(s) == 14:
        return f"{s[:2]}.{s[2:5]}.{s[5:8]}/{s[8:12]}-{s[12:14]}"
    return d


def _fmt_date_br(iso: str | None) -> str:
    if not iso:
        return ""
    parts = iso[:10].split("-")
    if len(parts) == 3:
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    return iso


def _write_header(ws: Worksheet, columns: list[tuple[str, str, str]]) -> None:
    ws.append([c[0] for c in columns])
    ws.row_dimensions[1].height = 25
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = CELL_BORDER


def _autofit(ws: Worksheet, columns: list[tuple[str, str, str]]) -> None:
    for col_idx, (header, _, kind) in enumerate(columns, 1):
        max_len = len(header)
        col_letter = get_column_letter(col_idx)
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            if kind == "currency" and isinstance(v, (int, float)):
                s = f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            else:
                s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)


def _build_lancamentos(ws: Worksheet, rows: list[dict]) -> None:
    _write_header(ws, LANC_COLUMNS)
    total = 0.0
    n = 0
    for r in rows:
        n += 1
        valor = float(r.get("valor_principal") or 0.0)
        total += valor
        ws.append([
            r.get("arquivo") or "",
            _fmt_cpf_cnpj(r.get("cnpj_destinatario")),
            valor,
            r.get("uf_favorecida") or "",
            _fmt_date_br(r.get("data_vencimento")),
            r.get("periodo_referencia") or "",
            r.get("no_controle") or "",
            r.get("criado_em") or "",
        ])
        row_n = n + 1
        ws.row_dimensions[row_n].height = 20
        for col_idx, (_, _, kind) in enumerate(LANC_COLUMNS, 1):
            cell = ws.cell(row=row_n, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = CENTER
            cell.border = CELL_BORDER
            if kind == "currency":
                cell.number_format = CURRENCY_FMT

    if n > 0:
        total_row = n + 2
        total_cell = ws.cell(row=total_row, column=1, value="TOTAL")
        total_cell.font = Font(name="Calibri", size=11, bold=True, color="1A1A1F")
        total_cell.alignment = CENTER
        total_cell.border = CELL_BORDER
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
        ws.cell(row=total_row, column=2).border = CELL_BORDER

        c_total = ws.cell(row=total_row, column=3, value=total)
        c_total.font = Font(name="Calibri", size=11, bold=True, color="059669")
        c_total.alignment = CENTER
        c_total.border = CELL_BORDER
        c_total.number_format = CURRENCY_FMT

        for col_idx in range(4, len(LANC_COLUMNS) + 1):
            c = ws.cell(row=total_row, column=col_idx)
            c.border = CELL_BORDER
            c.alignment = CENTER
        ws.row_dimensions[total_row].height = 22

    _autofit(ws, LANC_COLUMNS)
    ws.freeze_panes = "A2"


def _build_falhas(ws: Worksheet, rows: list[dict]) -> None:
    _write_header(ws, FALHAS_COLUMNS)
    for n, r in enumerate(rows, 1):
        ws.append([
            r.get("arquivo") or "",
            r.get("motivo") or "",
            r.get("criado_em") or "",
        ])
        row_n = n + 1
        ws.row_dimensions[row_n].height = 22
        for col_idx, (_, _, kind) in enumerate(FALHAS_COLUMNS, 1):
            cell = ws.cell(row=row_n, column=col_idx)
            cell.font = DATA_FONT
            cell.alignment = LEFT_WRAP if kind == "text_wrap" else CENTER
            cell.border = CELL_BORDER
    _autofit(ws, FALHAS_COLUMNS)
    ws.freeze_panes = "A2"


def build_xlsx_file(
    output_path: Path,
    lancamentos: Iterable[dict],
    falhas: Iterable[dict],
) -> None:
    wb = Workbook()
    ws_lanc = wb.active
    ws_lanc.title = "Lançamentos"
    _build_lancamentos(ws_lanc, list(lancamentos))

    ws_fail = wb.create_sheet("Falhas")
    _build_falhas(ws_fail, list(falhas))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
